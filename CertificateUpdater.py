# standard lib imports
import sys
from copy import copy
import datetime
# inport from openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


def cventDateFormat(dateString):
    """ Converts the dateString:
        from  YYYY-MM-DD HH:MM:SS
        to    MM/DD/YYYY HH:MM
    """
    localS = dateString.lstrip()
    y=localS[0:4]
    m=localS[5:7]
    d=localS[8:10]
    t=localS[11:16]
    cventDate=m+'/'+d+'/'+y+' '+t
    return cventDate


def copyColumn(sheetWS, fromColumn, toColumn):
    """ Copy the column from row 2 to max row. We do not want to change the header.
    """
    for row in range(2,sheetWS.max_row+1):
        sheetWS.cell(row=row, column=toColumn).value = sheetWS.cell(row=row, column=fromColumn).value


def setCompany(sheetWS):
    """ Copy the values in the organization column to the company column
    """
    fromColumn = findCellInRow(sheetWS[1], 'Organization')
    toColumn = findCellInRow(sheetWS[1], 'Company')
    copyColumn(sheetWS, fromColumn, toColumn)


def setParticipatedDict(participatedDict):
    participatedDict["Attended"] = "Yes"
    participatedDict["Yes"] = "Yes"

   
def translateValues(columnName, sheetWS, translationDict):
    """ This function finds the column that needs translating
        and then updates the values based on the dictionary.
    """
    column = findCellInRow(sheetWS[1], columnName)
    for row in range(2,sheetWS.max_row+1):
        cell = sheetWS.cell(row=row, column=column)
        if(cell.value is not None) and(cell.value != ''):
            cell.value = translationDict[cell.value]


def copyRow(fromRow, fromWS, toRow, toWS):
    """ This function copies a row of spreadsheet data
    """
    for column in range(1, fromWS.max_column+1):
        toCell = toWS.cell(row=toRow, column=column)
        fromCell = fromWS.cell(row=fromRow, column=column)
        toCell.value = fromCell.value
        toCell.data_type = fromCell.data_type
#        if fromCell.has_style:
#            toCell._style = copy(fromCell._style)
        if fromCell.hyperlink:
            toCell._hyperlink = copy(fromCell.hyperlink)
        if fromCell.comment:
            toCell.comment = copy(fromCell.comment)
    

def printSheetRow(row, rowType, fileName):
    """"Print a row of a spreadsheet"""
    print("\n", rowType, " from", fileName, ":")
    for cell in row:
        print(cell, " = ", cell.value)
    return 1


def findCellInRow(row, cellValue):
    """Seach a row to find a value and return the column index."""
    for cell in row:
        if cell.value == cellValue:
            return cell.col_idx
    return None


def setColumn(sheet, column, value):
    """ Set a column to a value."""
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row, column=CDAcountNumberColumn).value is not None:
           sheet.cell(row=row, column=column).value = value


def findCellRowInColumn(targetSheet, column, cellValue, extactCase=True):
    """Seach a column to find a value and return the row index.
       extactCase=True will match the strings extactly
       extactCase=False will match lower case versions of the strings"""

    # Need to work with spripped values
    cellStrippedValue = cellValue.strip()
    # Save some time and only convert the key string once
    if not extactCase:
        lowerCellValue = cellStrippedValue.lower()
    
    for row in range(1,targetSheet.max_row+1):
        if targetSheet.cell(row=row, column=column).value is not None:
            targetStrippedValue = targetSheet.cell(row=row, column=column).value.strip()
            if extactCase:
                if targetStrippedValue == cellStrippedValue:
                    return targetSheet.cell(row=row, column=column).row
            else:
                if targetStrippedValue.lower() == lowerCellValue:
                    return targetSheet.cell(row=row, column=column).row
        
    return None


def findRowInColumns(targetSheet, firstColumn, firstValue, secondColumn, secondValue, extactCase=True):
    """Match two colums to find a row. """
    # Save some time and only convert the key string once
    # print('In findRowInColumns extactCase = ', extactCase)
    # looks like everthing needs to be stripped
    fStrippedValue = firstValue.strip()
    sStrippedValue = secondValue.strip()

    if not extactCase:
        lowerFirstValue = fStrippedValue.lower()
        lowerSecondValue = sStrippedValue.lower()
        # print('lower case values: "'+lowerFirstValue+'", "'+lowerSecondValue+'"')
    
    for row in range(1,targetSheet.max_row+1):
        if targetSheet.cell(row=row, column=firstColumn).value is not None:
            fStrippedColumnValue = targetSheet.cell(row=row, column=firstColumn).value
            sStrippedColumnValue = targetSheet.cell(row=row, column=secondColumn).value
            if extactCase:
                if fStrippedColumnValue == fStrippedValue and \
                   sStrippedColumnValue == sStrippedValue:
                    return row
            else:
                if fStrippedColumnValue.lower() == lowerFirstValue and \
                   sStrippedColumnValue.lower() ==lowerSecondValue:
                    return row
        
    return None


def findRowIn3Columns(targetSheet, firstColumn, firstValue, secondColumn, secondValue, thirdColumn, thirdValue, extactCase=True):
    """Match three colums to find a row. """
    # Save some time and only convert the key string once
    # print('In findRowInColumns extactCase = ', extactCase)
    # looks like everthing needs to be stripped
    fStrippedValue = firstValue.strip()
    sStrippedValue = secondValue.strip()
    tStrippedValue = thirdValue.strip()

    if not extactCase:
        lowerFirstValue = fStrippedValue.lower()
        lowerSecondValue = sStrippedValue.lower()
        lowerThirdValue = tStrippedValue.lower()
        # print('lower case values: "'+lowerFirstValue+'", "'+lowerSecondValue+'"')
    
    for row in range(1,targetSheet.max_row+1):
        if targetSheet.cell(row=row, column=firstColumn).value is not None and\
           targetSheet.cell(row=row, column=secondColumn).value is not None and\
           targetSheet.cell(row=row, column=thirdColumn).value is not None:
            fStrippedColumnValue = targetSheet.cell(row=row, column=firstColumn).value
            sStrippedColumnValue = targetSheet.cell(row=row, column=secondColumn).value
            tStrippedColumnValue = targetSheet.cell(row=row, column=thirdColumn).value
            if extactCase:
                if fStrippedColumnValue == fStrippedValue and \
                   sStrippedColumnValue == sStrippedValue and \
                   tStrippedColumnValue == tStrippedValue:
                    return row
            else:
                if fStrippedColumnValue.lower() == lowerFirstValue and \
                   sStrippedColumnValue.lower() == lowerSecondValue and \
                   tStrippedColumnValue.lower() == lowerThirdValue:
                    return row
        
    return None


def findCellRowInColumnHeader(targetSheet, columnHeader, cellValue, extactCase=True):
    """ Find the column with the specified header and then find the row.
    """
    column = findCellInRow(row, columnHeader)
    if column is not None:
        return findCellRowInColumn(targetSheet, column, cellValue, extactCase)
    return None


def checkForNewCD (cdSheet, cdUpdateSheet):
    """ Confirm that all CDs in cdUpdatesSheet are in the cdSheet.
    """
    for row in range(2, cdUpdateSheet.max_row+1):
        cdRow = findCellRowInColumn(cdSheet, \
                               CDAcountNumberColumn, \
                               cdUpdateSheet.cell(row=row, \
                                                  column=CDUpdatesAcountNumberColumn).value)
        if cdRow is None:
            printSheetRow(cdUpdateSheet[row], "New CD found", row)
        else:
            # Found a match, update the row:
            cdSheet.cell(row=cdRow, column=CDCurrentValueColumn).value = \
                        cdUpdateSheet.cell(row=row, \
                                           column=CDUpdatesCurrentValueColumn).value
            cdSheet.cell(row=cdRow, column=CDAutoUpdatedColumn).value = datetime.datetime.now()
            cdSheet.cell(row=cdRow, column=CDTimesUpdatedColumn).value += 1
            # Has this CD renewed?
            maturity = cdSheet.cell(row=cdRow, column=CDCurrentMaturityColumn).value
            if maturity < datetime.datetime.now():
                print('CD has renewed: ', cdSheet.cell(row=cdRow, \
                      column=CDOwnerColumn).value, \
                      maturity, datetime.datetime.now(), \
                      cdUpdateSheet.cell(row=row, column=CDUpdatesAcountNumberColumn).value)


# path to files
path = "c:/Users/gja/Documents/py_test/"

# the files
CDFile = "Certificates Check.xlsx"
CDUpdatesFile = "Certificate Updates.xlsx"

print('loading ',CDFile)
CDWB = load_workbook(path+CDFile)
CDSheet = CDWB['Sheet1']
CDHeaderRow = CDSheet[1]
#printSheetRow(CDHeaderRow, "Headers", CDFile)

print('loading ',CDUpdatesFile)
CDUpdatesWB = load_workbook(path+CDUpdatesFile)
CDUpdatesSheet = CDUpdatesWB['Sheet1']
CDUpdatesHeaderRow = CDUpdatesSheet[1]
#printSheetRow(CDUpdatesHeaderRow, "Headers", CDUpdatesFile)

# Set header values
accountNumber = 'Account Number'
CDAcountNumberColumn = findCellInRow(CDSheet[1], accountNumber)
CDUpdatesAcountNumberColumn = findCellInRow(CDUpdatesSheet[1], accountNumber)

CDCurrentValue = 'Current Value'
CDUpdatesCurrentValue = 'Ending Balance'
CDCurrentValueColumn = findCellInRow(CDSheet[1], CDCurrentValue)
CDUpdatesCurrentValueColumn = findCellInRow(CDUpdatesSheet[1], CDUpdatesCurrentValue)

currentMaturity = 'Current Maturity'
CDCurrentMaturityColumn = findCellInRow(CDSheet[1], currentMaturity)

autoUpdated = 'Auto Updated'
CDAutoUpdatedColumn = findCellInRow(CDSheet[1], autoUpdated)

owner = 'Owner'
CDOwnerColumn = findCellInRow(CDSheet[1], owner)

timesUpdated = 'Times Updated'
CDTimesUpdatedColumn = findCellInRow(CDSheet[1], timesUpdated)

# Clear out Times updated
setColumn(CDSheet, CDTimesUpdatedColumn, 0)

# Check for new CDs
checkForNewCD (CDSheet, CDUpdatesSheet)

# Save the updated sheet
CDWB.save(path+CDFile)

# Exit
print('Done')

